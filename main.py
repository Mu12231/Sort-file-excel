from tkinter import *
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
from tkinter import messagebox

File = ''
File2 = ''
File3 = ' '


def excel_filter(path, file2):

    codes_coulmn = 'I'

    codes_coulmn2 = 'M'

    coulm_check1 = 'I'
    coulm_check2 = 'I'

    index = 6
    codes = []
    wb = load_workbook(path)
    ws1 = wb.active
    while (ws1[f'{codes_coulmn2}{index}'].value):
        if (ws1[f'{coulm_check1}{index}'].value == 0) and ((ws1[f'{coulm_check2}{index}'].value == 0)):
            codes.append(ws1[f'{codes_coulmn2}{index}'].value)

        index += 1

    wb2 = load_workbook(file2)
    ws2 = wb2.active
    ws2.auto_filter.ref = f"{codes_coulmn}:{codes_coulmn}"
    ws2.auto_filter.add_filter_column(0, codes, blank=True)
    ws2.auto_filter.add_sort_condition(f'{codes_coulmn}:{codes_coulmn}')
    wb.save(path)
    wb2.save(file2)
    messagebox.showinfo("Done", "Done")


def get_file(num):
    if num == 1:

        global File
        File = askopenfilename()
        label1.configure(text=File)
    elif num == 2:
        global File2
        File2 = askopenfilename()
        label2.configure(text=File2)
    else:
        global File3
        File3 = askopenfilename()
        label3.configure(text=File3)


def start():
    if not File or not File2 or not File3:
        return messagebox.showerror("Error", "Missing files")
    if (File == File2 or File == File3 or File2 == File3):
        return messagebox.showerror('Error', 'Duplicate Files')
    excel_filter(File, File2)
    excel_filter(File, File3)


File = ''
window = Tk()

window.title("M.C.S (Mu_1)")

window.geometry('400x150')


btn1 = Button(window, text="File1", command=lambda: get_file(1))
btn2 = Button(window, text="File2", command=lambda: get_file(2))
btn = Button(window, text="Sort", command=start)
btn3 = Button(window, text="File3", command=lambda: get_file(3))

label1 = Label(window, text="")


label1.grid(column=2, row=1)

label2 = Label(window, text="")


label2.grid(column=2, row=2)

label3 = Label(window, text="")

label3.grid(column=2, row=3)


btn1.grid(column=1, row=1)
btn2.grid(column=1, row=2)
btn.grid(column=1, row=4)
btn3.grid(column=1, row=3)

window.resizable(False, False)
window.mainloop()


# excel_filter()
